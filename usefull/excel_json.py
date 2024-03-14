# Запись Данных в Файл EXCEL

import openpyxl
import json

# book = openpyxl.Workbook()
# sheet = book.active
#
# sheet.cell(row=1, column=1).value = "Table's Name" # Строка (с 1)/ Колонка (с 1)
# sheet[2][0].value = 'header' # 'A2' Строка (с 1)/ Колонка (с 0)
# sheet['B2'] = 111
# sheet['C2'] = 'hello'
#
#
# excel_file = r'files/new_excel.xlsx'
# book.save(excel_file)
# book.close()

with open(r'files/movies.json') as file:
    data = json.load(file)
print(type(data))
# print(data)

book = openpyxl.Workbook()
sheet = book.active

sheet['A1'] = 'ID'
sheet['B1'] = 'TITLE'
sheet['C1'] = 'YEAR'
sheet['D1'] = 'GENRES'
sheet['E1'] = 'Director'
sheet['F1'] = 'Actors'

row = 2
for movie in data['movies']:
    sheet.cell(row=row, column=1).value = movie['id']
    sheet.cell(row=row, column=2).value = movie['title']
    sheet.cell(row=row, column=3).value = movie['year']
    sheet.cell(row=row, column=4).value = ', '.join(movie['genres']) # преобразую список в строку
    sheet.cell(row=row, column=5).value = movie['director']
    sheet.cell(row=row, column=6).value = movie['actors']
    row += 1

book.save(r'files/movies.xlsx')
book.close()
