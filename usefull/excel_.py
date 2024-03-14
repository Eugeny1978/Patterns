import openpyxl
from pprint import pprint

# Работа с Файлами Excel
# предварительно установи пакет pip install openpyxl
# Чтобы отображалось Автозаполнение (подсказки по атрибутам и модулям) удали папку помощи по адресу:
# C:\Program Files\JetBrains\PyCharm Community Edition 2023.2\plugins\python-ce\helpers\typeshed\stubs\openpyxl
# И перезагрузи IDE

file = r'files/Prices.xlsx'

pprint(locals())
pprint(dir(openpyxl))


book = openpyxl.load_workbook(file, read_only=True, data_only=True) # ==  openpyxl.open(file, read_only=True)
# data_only=True - покажет результаты вычислений формул в ячейках
sheet = book.active

print(sheet["C10"].value) # обратиться через Адрес ячейки
print(sheet[1][0].value) # обратиться через Индексы ячейки: [Строка - нач с 1] [Колонка - нач с 0]
print(sheet[10][2].value) # обратиться через Индексы ячейки: [Строка - нач с 1] [Колонка - нач с 0]

# last_row = sheet.max_row
# print(last_row)
#
# for row in range(7, last_row+1):
#     number = sheet[row][0].value
#     article = sheet[row][1].value
#     name = sheet[row][2].value
#     image = sheet[row][3].value
#     stock = sheet[row][4].value
#     price = sheet[row][5].value
#     order = sheet[row][6].value
#     print(number, article, name, image, stock, price, order, sep=' | ')
#
# cells = sheet['B7':'C85'] # диапазон (срез)
# for article, name in cells:
#     print(article.value, name.value, sep=' || ')

# Встроенный метод Итерации
# data_only=True - покажет результаты вычислений формул в ячейках.
# ожно указывать непосредственно прри итерации
# А сам лист открыть, не задавая явно это значение. По умолдчанию оно False.
for cell1, cell2, cell3, cell4, cell5, cell6, cell7 in sheet.iter_rows(min_row=7, max_row=20, min_col=1, max_col=7):
    print(cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value, cell7.value)

# Обратиться к конкретному листу
all_sheets = book.worksheets
print(all_sheets)
sheet_1 = all_sheets[1]
